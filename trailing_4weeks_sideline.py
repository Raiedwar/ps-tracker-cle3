"""
Trailing 4 Weeks - IB Problem Solve Sideline + Damageland Report
-----------------------------------------------------------------
Called by Launch PS Tracker.pyw — do not run directly.
Accepts optional argv[1] = base directory (defaults to script's own folder).
"""

import csv, re, sys, json
from datetime import datetime, timedelta, timezone
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# ── Paths & config ─────────────────────────────────────────────────────────────
BASE = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent
_cfg_file = BASE / "config.json"
_cfg = json.loads(_cfg_file.read_text()) if _cfg_file.exists() else {}

CSV_FOLDER  = BASE / "FCLM_CSVs"
OUTPUT_FILE = BASE / _cfg.get("report_filename", "PS_Tracker_Trailing4Weeks.xlsx")
ET_OFFSET     = timedelta(hours=-4)
WAREHOUSE_ID  = _cfg.get("warehouse_id", "CLE3")
PATH_NAME     = _cfg.get("path_name",    "IB Problem Solve")
SITE_LABEL    = _cfg.get("report_label", "Sideline & Damageland")

SL_FUNC_NAME       = _cfg.get("sl_func_name", "Stow to Prime PSolve")
SL_MIN_HRS_WEEK    = 1.0
SL_MIN_HRS_TOTAL   = 30.0
SL_MIN_UNITS_TOTAL = 2000

DL_FUNC_NAME       = "Damages"
DL_MIN_HRS_TOTAL   = 20.0
DL_MIN_JOBS_TOTAL  = 200
# ──────────────────────────────────────────────────────────────────────────────

FCLM_BASE = (
    "https://fclm-portal.amazon.com/reports/functionRollup?reportFormat=HTML"
    "&warehouseId=CLE3&processId=1002980&spanType=Week"
    "&maxIntradayDays=1&startHourIntraday=0&startMinuteIntraday=0"
    "&endHourIntraday=0&endMinuteIntraday=0"
    "&startHourIntraday1=7&startMinuteIntraday1=0"
    "&startHourIntraday2=17&startMinuteIntraday2=2"
    "&startHourIntraday3=18&startMinuteIntraday3=0"
    "&startHourIntraday4=4&startMinuteIntraday4=2"
)

def parse_week_label(ts_str):
    dt_utc = datetime.strptime(ts_str, "%Y%m%d%H%M%S").replace(tzinfo=timezone.utc)
    dt_et  = dt_utc + ET_OFFSET
    return dt_et.strftime("%b %d"), dt_et

def fclm_url(dt_et):
    wk_start = dt_et.strftime("%Y/%m/%d").replace("/", "%2F")
    wk_end   = (dt_et + timedelta(days=6)).strftime("%Y/%m/%d").replace("/", "%2F")
    return f"{FCLM_BASE}&startDateDay={wk_end}&startDateWeek={wk_start}"

def extract_sideline(filepath):
    total_units = 0; hrs_map = {}; assoc = {}
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if (row["Function Name"] == SL_FUNC_NAME
                    and row["Job Action"] == "SidelineApp"
                    and row["Unit Type"] == "EACH"
                    and row["Size"] == "Total"):
                eid = row["Employee Id"]; name = row["Name"]
                u = float(row["Units"]); h = float(row["Paid Hours-Total(function,employee)"])
                total_units += u; hrs_map[eid] = h
                if eid not in assoc: assoc[eid] = [name, 0.0, h]
                assoc[eid][1] += u
    total_hrs = sum(hrs_map.values())
    uph = round(total_units / total_hrs, 2) if total_hrs else 0
    out = {}
    for eid, (name, u, h) in assoc.items():
        w_uph = round(u / h, 2) if h >= SL_MIN_HRS_WEEK else None
        out[eid] = (name, int(u), round(h, 2), w_uph)
    return int(total_units), round(total_hrs, 2), uph, out

def extract_damageland(filepath):
    out = {}
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if (row["Function Name"] == DL_FUNC_NAME
                    and row["Job Action"] == "ItemGraded"
                    and row["Size"] == "Total"):
                eid = row["Employee Id"]; name = row["Name"]
                j = float(row["Jobs"]); h = float(row["Paid Hours-Total(function,employee)"])
                jph = float(row["JPH"]) if row["JPH"] else 0
                out[eid] = (name, int(j), round(h, 2), round(jph, 2))
    total_jobs = sum(v[1] for v in out.values())
    total_hrs  = sum(v[2] for v in out.values())
    avg_jph    = round(total_jobs / total_hrs, 2) if total_hrs else 0
    return int(total_jobs), round(total_hrs, 2), avg_jph, out

def single_week_sl_ranks(sl_map):
    """Rank all associates within a single week by UPH. Rank 1 = highest UPH."""
    ranked = sorted(
        [(eid, v[3]) for eid, v in sl_map.items() if v[3] is not None],
        key=lambda x: x[1], reverse=True
    )
    return {eid: i + 1 for i, (eid, _) in enumerate(ranked)}

def single_week_dl_ranks(dl_map):
    """Rank all associates within a single week by JPH. Rank 1 = highest JPH."""
    ranked = sorted(
        [(eid, v[3]) for eid, v in dl_map.items() if v[3] > 0],
        key=lambda x: x[1], reverse=True
    )
    return {eid: i + 1 for i, (eid, _) in enumerate(ranked)}

def percentile_ranks(values):
    n = len(values)
    if n == 0: return {}
    sorted_keys = [k for k, v in sorted(values, key=lambda x: x[1])]
    return {k: i / (n - 1) if n > 1 else 1.0 for i, k in enumerate(sorted_keys)}

def build_sl_ranking(weeks_subset, five_weeks_maps=None):
    """Build sideline associate data + overall ranking across weeks.
    five_weeks_maps: list of 5 sl_map dicts (week0 through week4); index 0 = hidden baseline."""
    all_eids = {}
    for _, _, _, _, _, sl_map, *_ in weeks_subset:
        for eid, (name, *_) in sl_map.items(): all_eids[eid] = name

    data = {}
    for eid, name in all_eids.items():
        wk_units = []; wk_uph = []; wk_hours = []
        for _, _, _, _, _, sl_map, *_ in weeks_subset:
            e = sl_map.get(eid)
            if e: wk_units.append(e[1]); wk_uph.append(e[3]); wk_hours.append(e[2])
            else: wk_units.append(None); wk_uph.append(None); wk_hours.append(None)
        valid_uphs = [u for u in wk_uph if u is not None]
        if not valid_uphs: continue
        total_units = sum(u for u in wk_units if u is not None)
        total_hours = sum(h for h in wk_hours if h is not None)
        if total_hours < SL_MIN_HRS_TOTAL: continue
        if total_units < SL_MIN_UNITS_TOTAL: continue
        avg_uph = round(sum(valid_uphs) / len(valid_uphs), 2)

        # Per-week UPH delta: compare this week's UPH to the prior week's UPH
        # five_weeks_maps[0] = hidden week0, [1..4] = displayed weeks 1-4
        wk_rate_delta = []
        if five_weeks_maps:
            for j in range(4):
                curr_uph = (five_weeks_maps[j + 1].get(eid) or (None,None,None,None))[3]
                prev_uph = (five_weeks_maps[j].get(eid)     or (None,None,None,None))[3]
                if curr_uph is not None and prev_uph is not None:
                    wk_rate_delta.append(round(curr_uph - prev_uph, 2))
                else:
                    wk_rate_delta.append(None)
        else:
            wk_rate_delta = [None, None, None, None]

        data[eid] = {"name": name, "wk_units": wk_units, "wk_uph": wk_uph,
                     "wk_hours": wk_hours, "wk_rate_delta": wk_rate_delta,
                     "total_units": total_units,
                     "total_hours": round(total_hours, 2), "avg_uph": avg_uph}

    eids    = list(data.keys())
    u_pct   = percentile_ranks([(e, data[e]["total_units"]) for e in eids])
    uph_pct = percentile_ranks([(e, data[e]["avg_uph"])     for e in eids])
    for eid in eids:
        data[eid]["master"] = (u_pct[eid] * 0.5) + (uph_pct[eid] * 0.5)
    ranked = sorted(eids, key=lambda e: data[e]["master"], reverse=True)
    return data, ranked

def build_dl_ranking(weeks_subset, five_weeks_maps=None):
    """Build damageland associate data + overall ranking across weeks.
    five_weeks_maps: list of 5 dl_map dicts (week0 through week4); index 0 = hidden baseline."""
    all_eids = {}
    for _, _, _, _, _, _, _, _, _, dl_map, *_ in weeks_subset:
        for eid, (name, *_) in dl_map.items(): all_eids[eid] = name

    data = {}
    for eid, name in all_eids.items():
        wk_jobs = []; wk_jph = []; wk_hours = []
        for _, _, _, _, _, _, _, _, _, dl_map, *_ in weeks_subset:
            e = dl_map.get(eid)
            if e: wk_jobs.append(e[1]); wk_jph.append(e[3]); wk_hours.append(e[2])
            else: wk_jobs.append(None); wk_jph.append(None); wk_hours.append(None)
        valid_jphs = [j for j in wk_jph if j is not None]
        if not valid_jphs: continue
        total_jobs  = sum(j for j in wk_jobs if j is not None)
        total_hours = sum(h for h in wk_hours if h is not None)
        if total_hours < DL_MIN_HRS_TOTAL: continue
        if total_jobs  < DL_MIN_JOBS_TOTAL: continue
        avg_jph = round(sum(valid_jphs) / len(valid_jphs), 2)

        wk_rate_delta = []
        if five_weeks_maps:
            for j in range(4):
                curr_jph = (five_weeks_maps[j + 1].get(eid) or (None,None,None,None))[3]
                prev_jph = (five_weeks_maps[j].get(eid)     or (None,None,None,None))[3]
                if curr_jph is not None and prev_jph is not None:
                    wk_rate_delta.append(round(curr_jph - prev_jph, 2))
                else:
                    wk_rate_delta.append(None)
        else:
            wk_rate_delta = [None, None, None, None]

        data[eid] = {"name": name, "wk_jobs": wk_jobs, "wk_jph": wk_jph,
                     "wk_hours": wk_hours, "wk_rate_delta": wk_rate_delta,
                     "total_jobs": total_jobs,
                     "total_hours": round(total_hours, 2), "avg_jph": avg_jph}

    ranked = sorted(data.keys(), key=lambda e: data[e]["avg_jph"], reverse=True)
    return data, ranked

def movement_label(prev_avg_rate, curr_avg_rate):
    """Column C: 4-week avg UPH/JPH change vs previous 4-week window."""
    if prev_avg_rate is None:
        return "NEW", "2E75B6"
    diff = round(curr_avg_rate - prev_avg_rate, 2)
    if diff > 0:  return f"+{diff}", "375623"
    if diff < 0:  return f"{diff}", "C00000"
    return "—", "808080"

def rate_delta_label(delta):
    """Per-week UPH/JPH change vs previous week."""
    if delta is None:
        return "", "808080"
    if delta > 0:  return f"+{delta}", "375623"
    if delta < 0:  return f"{delta}", "C00000"
    return "—", "808080"

def style(cell, bold=False, bg=None, fg="000000", align="center", size=11, border=False, num_fmt=None):
    cell.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if border:
        s = Side(style="thin", color="BFBFBF")
        cell.border = Border(left=s, right=s, top=s, bottom=s)
    if num_fmt:
        cell.number_format = num_fmt

def main():
    csv_files = sorted(
        f for f in CSV_FOLDER.glob(f"functionRollupReport-{WAREHOUSE_ID}-{PATH_NAME}-Week-*.csv")
        if re.search(r"-Week-(\d{14})-(\d{14})\.csv$", f.name)
    )
    if not csv_files:
        print(f"No CSVs found in:\n  {CSV_FOLDER}")
        return

    all_weeks = []
    for f in csv_files:
        m = re.search(r"-Week-(\d{14})-(\d{14})\.csv$", f.name)
        lbl, dt = parse_week_label(m.group(1))
        sl_u, sl_h, sl_uph, sl_map = extract_sideline(f)
        dl_j, dl_h, dl_jph, dl_map = extract_damageland(f)
        url = fclm_url(dt)
        all_weeks.append((dt, lbl, sl_u, sl_h, sl_uph, sl_map, dl_j, dl_h, dl_jph, dl_map, url))

    all_weeks.sort(key=lambda x: x[0])

    # Displayed weeks = last 4; week0 = the week before (hidden, used as rank baseline)
    weeks = all_weeks[-4:]
    week0 = all_weeks[-5] if len(all_weeks) >= 5 else None

    # Previous 4-week window for overall rank movement (column C)
    prev_weeks = all_weeks[-5:-1] if len(all_weeks) >= 5 else None

    print("\nTrailing 4 weeks:\n")
    print(f"  {'Week':<12} {'SL Units':>10}  {'SL UPH':>8}  {'DL Jobs':>9}  {'DL JPH':>8}")
    print(f"  {'-'*54}")
    for _, lbl, su, sh, suph, _, dj, dh, djph, _, _ in weeks:
        print(f"  {lbl:<12} {su:>10,}  {suph:>8.2f}  {dj:>9,}  {djph:>8.2f}")
    if week0:
        print(f"\n  Hidden week (rank baseline): Wk of {week0[1]}")

    # ── Build per-week rank maps (weeks 0–4) for the Δ rank columns ───────────
    # Index 0 = hidden week0, indices 1-4 = displayed weeks
    five_weeks_sl = [week0[5] if week0 else {}] + [w[5] for w in weeks]
    five_weeks_dl = [week0[9] if week0 else {}] + [w[9] for w in weeks]
    sl_per_week_ranks = [single_week_sl_ranks(m) for m in five_weeks_sl]
    dl_per_week_ranks = [single_week_dl_ranks(m) for m in five_weeks_dl]

    # ── Current overall rankings (trailing 4 weeks) ───────────────────────────
    sl_data, sl_ranked = build_sl_ranking(weeks, five_weeks_sl if week0 else None)
    dl_data, dl_ranked = build_dl_ranking(weeks, five_weeks_dl if week0 else None)
    sl_n = len(sl_ranked)
    sl_top  = max(1, round(sl_n * 0.20))
    sl_bot  = max(1, round(sl_n * 0.20))
    sl_warn = max(1, round(sl_n * 0.25))
    dl_n = len(dl_ranked)
    dl_top  = max(1, round(dl_n * 0.20))
    dl_bot  = max(1, round(dl_n * 0.20))
    dl_warn = max(1, round(dl_n * 0.25))

    # ── Previous window rankings for column C overall movement ────────────────
    if prev_weeks:
        sl_prev_data, _ = build_sl_ranking(prev_weeks)
        dl_prev_data, _ = build_dl_ranking(prev_weeks)
        sl_prev_avg_rate = {eid: sl_prev_data[eid]["avg_uph"] for eid in sl_prev_data}
        dl_prev_avg_rate = {eid: dl_prev_data[eid]["avg_jph"] for eid in dl_prev_data}
        print(f"\n  Movement tracking: ON")
    else:
        sl_prev_avg_rate = {}; dl_prev_avg_rate = {}
        print(f"\n  Movement tracking: OFF  (need 5+ weeks for UPH deltas)")

    print(f"\n  Sideline: {sl_n} associates  |  Damageland: {dl_n} associates")

    # ── Excel ─────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trailing 4 Weeks"

    NAVY    = "1F3864"; BLUE = "2E75B6"; WHITE = "FFFFFF"
    DGRAY   = "404040"; LGRAY = "F2F2F2"; GOLD = "FFD966"
    GRN_BG  = "375623"; GRN_FG = "FFFFFF"
    RED_BG  = "C00000"; RED_FG = "FFFFFF"
    GOLD_BG = "FFD966"; GOLD_FG = "404040"
    WK_COLS  = ["4472C4","ED7D31","70AD47","7030A0"]
    WK_LIGHT = ["D9E1F2","FCE4D6","E2EFDA","EAD1F2"]

    # ── Column layout ─────────────────────────────────────────────────────────
    # B: Name | C: Δ Avg UPH
    # Each week = 4 cols: Units | Rate | Δ UPH/JPH | Hours
    #   Wk1 → D, E, F, G
    #   Wk2 → H, I, J, K
    #   Wk3 → L, M, N, O
    #   Wk4 → P, Q, R, S
    # T: 4-Wk Total | U: Total Hours | V: Avg Rate | W: Score | X: margin
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 9
    for col in ["D","E","G","H","I","K","L","M","O","P","Q","S"]:
        ws.column_dimensions[col].width = 10
    for col in ["F","J","N","R"]:
        ws.column_dimensions[col].width = 8   # Δ UPH/JPH
    ws.column_dimensions["T"].width = 14
    ws.column_dimensions["U"].width = 12
    ws.column_dimensions["V"].width = 12
    ws.column_dimensions["W"].width = 10
    ws.column_dimensions["X"].width = 2

    # wk_quads: (Units, Rate, Δ Rank, Hours) per displayed week
    wk_quads    = [("D","E","F","G"), ("H","I","J","K"), ("L","M","N","O"), ("P","Q","R","S")]
    week_labels = [w[1] for w in weeks]

    # ── Frozen column header row (row 1) — week dates included so labels stay clear while scrolling ──
    ws.row_dimensions[1].height = 22
    hdr_cols = ["B", "C"]
    hdr_vals = ["Associate", "Δ Avg UPH"]
    for wk_lbl, (c1, c2, c3, c4) in zip(week_labels, wk_quads):
        hdr_cols += [c1, c2, c3, c4]
        hdr_vals += [f"{wk_lbl} Units", f"{wk_lbl} UPH", f"{wk_lbl} Δ UPH", f"{wk_lbl} Hours"]
    hdr_cols += ["T", "U", "V", "W"]
    hdr_vals += ["Total Units", "Total Hours", "Avg UPH/JPH", "Score"]
    for col, h in zip(hdr_cols, hdr_vals):
        c = ws[f"{col}1"]; c.value = h
        style(c, bold=True, bg="2E75B6", fg="FFFFFF", size=10, border=True)

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("B2:W2")
    ws["B2"] = "IB Problem Solve — Sideline & Damageland Performance"
    style(ws["B2"], bold=True, bg=NAVY, fg=WHITE, size=14)
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:W3")
    ws["B3"] = (f"{WAREHOUSE_ID}  |  {SITE_LABEL}  |  Trailing 4 Weeks  |  "
                f"Generated: {datetime.now().strftime('%m/%d/%Y')}")
    style(ws["B3"], bg=BLUE, fg=WHITE, size=10)
    ws.row_dimensions[3].height = 18

    # ── Weekly summary table ──────────────────────────────────────────────────
    ws.row_dimensions[5].height = 20
    for col, hdr in zip(["B","C","D","E","F","G","H"],
                         ["Week Of","","SL Units","SL Hours","SL UPH","DL Jobs","DL Hours"]):
        c = ws[f"{col}5"]; c.value = hdr
        style(c, bold=True, bg=BLUE, fg=WHITE, size=11, border=True)
    ws["H5"] = "DL JPH"
    style(ws["H5"], bold=True, bg=BLUE, fg=WHITE, size=11, border=True)

    for i, (_, lbl, su, sh, suph, _, dj, dh, djph, _, _) in enumerate(weeks):
        row = 6 + i; ws.row_dimensions[row].height = 18
        ws[f"B{row}"] = f"Wk of {lbl}"
        ws[f"D{row}"] = su; ws[f"E{row}"] = sh; ws[f"F{row}"] = suph
        ws[f"G{row}"] = dj; ws[f"H{row}"] = dh; ws[f"I{row}"] = djph
        style(ws[f"B{row}"], bold=True, bg=WK_COLS[i], fg=WHITE, size=11, border=True)
        style(ws[f"C{row}"], bg=WK_LIGHT[i], fg=DGRAY, size=11, border=True)
        style(ws[f"D{row}"], bg=WK_LIGHT[i], fg=DGRAY, size=11, border=True, num_fmt="#,##0")
        style(ws[f"E{row}"], bg=WK_LIGHT[i], fg=DGRAY, size=11, border=True, num_fmt="#,##0.00")
        style(ws[f"F{row}"], bg=WK_LIGHT[i], fg=DGRAY, size=11, border=True, num_fmt="#,##0.00")
        style(ws[f"G{row}"], bg=WK_LIGHT[i], fg=DGRAY, size=11, border=True, num_fmt="#,##0")
        style(ws[f"H{row}"], bg=WK_LIGHT[i], fg=DGRAY, size=11, border=True, num_fmt="#,##0.00")
        style(ws[f"I{row}"], bg=WK_LIGHT[i], fg=DGRAY, size=11, border=True, num_fmt="#,##0.00")

    tr = 10; ws.row_dimensions[tr].height = 20
    ws[f"B{tr}"] = "4-Wk Total"
    ws[f"D{tr}"] = sum(w[2] for w in weeks)
    ws[f"E{tr}"] = round(sum(w[3] for w in weeks), 2)
    ws[f"F{tr}"] = round(sum(w[2] for w in weeks)/sum(w[3] for w in weeks), 2)
    ws[f"G{tr}"] = sum(w[6] for w in weeks)
    ws[f"H{tr}"] = round(sum(w[7] for w in weeks), 2)
    ws[f"I{tr}"] = round(sum(w[6] for w in weeks)/sum(w[7] for w in weeks), 2)
    for col in ["B","C","D","E","F","G","H","I"]:
        style(ws[f"{col}{tr}"], bold=True, bg=GOLD, fg=DGRAY, size=11, border=True)
    for col, fmt in [("D","#,##0"),("E","#,##0.00"),("F","#,##0.00"),
                     ("G","#,##0"),("H","#,##0.00"),("I","#,##0.00")]:
        ws[f"{col}{tr}"].number_format = fmt

    # ── Charts ────────────────────────────────────────────────────────────────
    for i, (_, lbl, su, _, suph, _, dj, _, djph, _, _) in enumerate(reversed(weeks)):
        ws.cell(row=14+i, column=2, value=f"Wk {lbl}")
        ws.cell(row=14+i, column=3, value=su)
        ws.cell(row=14+i, column=4, value=suph)
        ws.cell(row=14+i, column=5, value=dj)
        ws.cell(row=14+i, column=6, value=djph)

    cats = Reference(ws, min_col=2, min_row=14, max_row=17)
    def bar_chart(title, x_title, data_col, color, anchor):
        c = BarChart(); c.type = "bar"; c.title = title
        c.x_axis.title = x_title; c.y_axis.title = "Week"
        c.style = 10; c.width = 18; c.height = 11; c.legend = None
        c.add_data(Reference(ws, min_col=data_col, min_row=13, max_row=17), titles_from_data=True)
        c.set_categories(cats)
        c.series[0].graphicalProperties.solidFill = color
        ws.add_chart(c, anchor)

    bar_chart("Sideline Units",    "Units",     3, "2E75B6", "B13")
    bar_chart("Sideline UPH",      "UPH",       4, "1F3864", "D13")
    bar_chart("Damageland Jobs",   "Jobs",      5, "ED7D31", "B36")
    bar_chart("Damageland JPH",    "Jobs/Hour", 6, "7030A0", "D36")

    # ── Associate section writer ───────────────────────────────────────────────
    week_urls = [w[10] for w in weeks]

    def write_section(start, title, subtitle, ranked_eids, data, n, top_cut, bot_cut, warn_cut,
                      prev_avg_rate, col1_lbl, wk1_lbl, wk2_lbl,
                      get_wk1, get_wk2, get_wk_rank_delta, get_wk_hrs,
                      tot1_lbl, tot2_lbl, tot3_lbl,
                      get_tot1, get_tot2, get_tot3, rate_lbl="Δ UPH", score_key=None):

        ws.merge_cells(f"B{start}:W{start}")
        ws[f"B{start}"] = title
        style(ws[f"B{start}"], bold=True, bg=NAVY, fg=WHITE, size=13)
        ws.row_dimensions[start].height = 26

        leg = start + 1
        ws.merge_cells(f"B{leg}:W{leg}")
        ws[f"B{leg}"] = subtitle
        style(ws[f"B{leg}"], bg=BLUE, fg=WHITE, size=9)
        ws.row_dimensions[leg].height = 16

        grp = start + 2; hdr = start + 3
        ws.row_dimensions[grp].height = 16; ws.row_dimensions[hdr].height = 20

        # Week group headers spanning all 4 columns
        for idx, (c1, c2, c3, c4) in enumerate(wk_quads):
            ws.merge_cells(f"{c1}{grp}:{c4}{grp}")
            ws[f"{c1}{grp}"] = f"Wk {week_labels[idx]}"
            style(ws[f"{c1}{grp}"], bold=True, bg=WK_COLS[idx], fg=WHITE, size=10, border=True)

        for col, val in [("B",""),("C",""),("T","4-Wk Total"),("U",""),("V",""),("W","")]:
            ws[f"{col}{grp}"] = val
            style(ws[f"{col}{grp}"], bold=True, bg=NAVY, fg=WHITE, size=10, border=True)

        sub_cols = ["B","C"] + [c for t in wk_quads for c in t] + ["T","U","V","W"]
        sub_hdrs = [col1_lbl, "Δ Avg UPH"] + [wk1_lbl, wk2_lbl, rate_lbl, "Hours"] * 4 + \
                   [tot1_lbl, tot2_lbl, tot3_lbl, "Score" if score_key else ""]
        for col, h in zip(sub_cols, sub_hdrs):
            c = ws[f"{col}{hdr}"]; c.value = h
            style(c, bold=True, bg=BLUE, fg=WHITE, size=10, border=True)

        dr = start + 4
        for i, eid in enumerate(ranked_eids):
            d = data[eid]; r = dr + i
            ws.row_dimensions[r].height = 16

            if i < top_cut:          bg, fg = GRN_BG, GRN_FG
            elif i >= n - bot_cut:   bg, fg = RED_BG, RED_FG
            elif i >= n - warn_cut:  bg, fg = GOLD_BG, GOLD_FG
            else:                    bg = WHITE if i % 2 == 0 else LGRAY; fg = DGRAY

            # Name — clickable Engage link
            url = f"https://atoz.amazon.work/engage/conversation-hub/employee/{eid}/overview"
            nc = ws[f"B{r}"]; nc.value = d["name"]; nc.hyperlink = url
            style(nc, bg=bg, fg=fg, align="left", size=10, border=True)
            nc.font = Font(name="Calibri", size=10, color=fg, underline="single", bold=False)

            # Column C — avg UPH/JPH change vs previous 4-week window
            lbl_mv, color_mv = movement_label(prev_avg_rate.get(eid), data[eid].get("avg_uph") or data[eid].get("avg_jph"))
            mv_cell = ws[f"C{r}"]; mv_cell.value = lbl_mv
            style(mv_cell, bold=True, bg=bg, fg=color_mv, size=10, border=True)

            # Weekly columns — Units | Rate | Δ Rank | Hours
            for j, (c1, c2, c3, c4) in enumerate(wk_quads):
                v1  = get_wk1(d, j)
                v2  = get_wk2(d, j)
                dv  = get_wk_rank_delta(d, j)
                v4  = get_wk_hrs(d, j)
                cb  = bg if (i < top_cut or i >= n - warn_cut) else WK_LIGHT[j]
                cf  = fg if (i < top_cut or i >= n - warn_cut) else DGRAY
                wurl = week_urls[j]

                # Units
                c = ws[f"{c1}{r}"]; c.value = v1 if v1 is not None else ""
                style(c, bg=cb, fg=cf, size=10, border=True, num_fmt="#,##0" if v1 is not None else None)
                if v1 is not None: c.hyperlink = wurl; c.font = Font(name="Calibri", size=10, color=cf, underline="single")

                # Rate (UPH or JPH)
                c = ws[f"{c2}{r}"]; c.value = v2 if v2 is not None else ""
                style(c, bg=bg, fg=fg, size=10, border=True, num_fmt="#,##0.00" if v2 is not None else None)
                if v2 is not None: c.hyperlink = wurl; c.font = Font(name="Calibri", size=10, color=fg, underline="single")

                # Δ UPH/JPH — week-over-week rate change
                dlbl, dcolor = rate_delta_label(dv)
                dc = ws[f"{c3}{r}"]; dc.value = dlbl
                style(dc, bold=True, bg=bg, fg=dcolor, size=9, border=True)

                # Hours
                c = ws[f"{c4}{r}"]; c.value = v4 if v4 is not None else ""
                style(c, bg=bg, fg=fg, size=10, border=True, num_fmt="#,##0.00" if v4 is not None else None)
                if v4 is not None: c.hyperlink = wurl; c.font = Font(name="Calibri", size=10, color=fg, underline="single")

            # Totals
            c = ws[f"T{r}"]; c.value = get_tot1(d)
            style(c, bold=True, bg=bg, fg=fg, size=10, border=True, num_fmt="#,##0")
            c = ws[f"U{r}"]; c.value = get_tot2(d)
            style(c, bold=True, bg=bg, fg=fg, size=10, border=True, num_fmt="#,##0.00")
            c = ws[f"V{r}"]; c.value = get_tot3(d)
            style(c, bold=True, bg=bg, fg=fg, size=10, border=True, num_fmt="#,##0.00")

            # Score
            if score_key:
                sc = round(d[score_key] * 100, 1)
                c = ws[f"W{r}"]; c.value = sc
                style(c, bold=True, bg=bg, fg=fg, size=10, border=True, num_fmt="0.0")

    # ── Sideline ──────────────────────────────────────────────────────────────
    write_section(
        start       = 60,
        title       = "Sideline Associate Performance — Trailing 4 Weeks",
        subtitle    = (f"Ranked: 50% Total Units + 50% Avg UPH  |  "
                       f"Top 20% ({sl_top}) = Green  |  Warning = Yellow  |  Bottom 20% ({sl_bot}) = Red  |  "
                       f"Min {SL_MIN_HRS_TOTAL}h & {SL_MIN_UNITS_TOTAL:,} units  |  "
                       f"Δ Avg UPH = avg UPH vs previous 4-week window (+green, -red)  |  "
                       f"Δ UPH = week-over-week UPH change (+green, -red)"),
        ranked_eids = sl_ranked, data = sl_data, n = sl_n,
        top_cut = sl_top, bot_cut = sl_bot, warn_cut = sl_warn,
        prev_avg_rate    = sl_prev_avg_rate,
        col1_lbl         = "Associate",
        wk1_lbl          = "Units", wk2_lbl = "UPH",
        get_wk1          = lambda d, j: d["wk_units"][j],
        get_wk2          = lambda d, j: d["wk_uph"][j],
        get_wk_rank_delta= lambda d, j: d["wk_rate_delta"][j],
        rate_lbl         = "Δ UPH",
        get_wk_hrs       = lambda d, j: d["wk_hours"][j],
        tot1_lbl         = "Total Units", tot2_lbl = "Total Hours", tot3_lbl = "Avg UPH",
        get_tot1         = lambda d: d["total_units"],
        get_tot2         = lambda d: d["total_hours"],
        get_tot3         = lambda d: d["avg_uph"],
        score_key        = "master",
    )

    # ── Damageland ────────────────────────────────────────────────────────────
    dl_start = 60 + 4 + sl_n + 5
    write_section(
        start       = dl_start,
        title       = "Damageland Associate Performance — Trailing 4 Weeks",
        subtitle    = (f"Ranked by Avg JPH  |  "
                       f"Top 20% ({dl_top}) = Green  |  Warning = Yellow  |  Bottom 20% ({dl_bot}) = Red  |  "
                       f"Min {DL_MIN_HRS_TOTAL}h & {DL_MIN_JOBS_TOTAL} jobs  |  "
                       f"Δ Avg JPH = avg JPH vs previous 4-week window (+green, -red)  |  "
                       f"Δ JPH = week-over-week JPH change (+green, -red)"),
        ranked_eids = dl_ranked, data = dl_data, n = dl_n,
        top_cut = dl_top, bot_cut = dl_bot, warn_cut = dl_warn,
        prev_avg_rate    = dl_prev_avg_rate,
        col1_lbl         = "Associate",
        wk1_lbl          = "Jobs", wk2_lbl = "JPH",
        get_wk1          = lambda d, j: d["wk_jobs"][j],
        get_wk2          = lambda d, j: d["wk_jph"][j],
        get_wk_rank_delta= lambda d, j: d["wk_rate_delta"][j],
        rate_lbl         = "Δ JPH",
        get_wk_hrs       = lambda d, j: d["wk_hours"][j],
        tot1_lbl         = "Total Jobs", tot2_lbl = "Total Hours", tot3_lbl = "Avg JPH",
        get_tot1         = lambda d: d["total_jobs"],
        get_tot2         = lambda d: d["total_hours"],
        get_tot3         = lambda d: d["avg_jph"],
        score_key        = None,
    )

    ws.freeze_panes = "A2"

    lock = OUTPUT_FILE.parent / f"~${OUTPUT_FILE.name}"
    if lock.exists():
        print(f"\nERROR: Excel has the report open. Close it and run again.")
        raise PermissionError("File locked by Excel")
    try:
        wb.save(OUTPUT_FILE)
        print(f"\nSaved to: {OUTPUT_FILE}")
    except PermissionError:
        print(f"\nERROR: Cannot save — close the Excel report and run again.")
        raise

if __name__ == "__main__":
    main()
